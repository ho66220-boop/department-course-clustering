# -*- coding: utf-8 -*-
"""Expert card-sorting analysis: consensus clustering and agreement with the
algorithmic (IDF / binary) clusters.

Input: the completed master workbook with 14 raters' group numbers in the
평가 sheet (rows 6-30 = 25 departments, columns C-P = raters). The file
contains real respondent names, so it is gitignored; only aggregate,
department-level outputs (co-occurrence, consensus clusters, agreement) are
written to results/.

Outputs (results/figures/keep_for_report, results/tables/keep_for_report):
- cardsort_cooccurrence_heatmap.png : 25x25 co-classification, consensus order
- cardsort_consensus_dendrogram.png : dendrogram from 1 - co-occurrence
- cardsort_agreement_bars.png       : consensus vs IDF/binary ARI at k=3,4
- cardsort_cooccur_vs_idf_scatter.png : 300 pairs, co-occurrence vs IDF cosine
- cardsort_consensus_assignments.csv, cardsort_agreement.csv,
  cardsort_department_ambiguity.csv

Run:
    python src/build_card_sorting_analysis.py
"""
from __future__ import annotations

import itertools
import pathlib

import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from scipy.cluster.hierarchy import dendrogram, fcluster, leaves_list, linkage
from scipy.spatial.distance import pdist, squareform
from sklearn.metrics import adjusted_rand_score, normalized_mutual_info_score

BASE = pathlib.Path(__file__).resolve().parents[1]
CANDIDATE_INPUTS = [
    BASE / "data" / "raw" / "card_sorting_responses" / "CardSorting_master_조사완료.xlsx",
    BASE / "CardSorting_master_조사완료.xlsx",
]
MATRIX = BASE / "data" / "processed" / "department_course_matrix_binary.csv"
REPORT_TABLES = BASE / "results" / "tables" / "keep_for_report"
REPORT_FIGURES = BASE / "results" / "figures" / "keep_for_report"

METADATA = {"department_id", "department_name", "department_name_ko", "broad_field", "selected_reason"}
RATER_FIRST_COL, RATER_LAST_COL = 3, 16   # 평가 columns C..P (14 raters)
DEPT_FIRST_ROW, DEPT_LAST_ROW = 6, 30     # 25 departments
NAME_COL = 2
K_PRIMARY = 4


def setup_style() -> None:
    matplotlib.rcParams["font.family"] = ["Malgun Gothic", "DejaVu Sans"]
    matplotlib.rcParams["axes.unicode_minus"] = False


def find_input() -> pathlib.Path:
    for path in CANDIDATE_INPUTS:
        if path.exists():
            return path
    raise FileNotFoundError(
        "Completed card-sorting workbook not found. Looked in:\n  "
        + "\n  ".join(str(p) for p in CANDIDATE_INPUTS)
    )


def read_responses(path: pathlib.Path) -> pd.DataFrame:
    ws = openpyxl.load_workbook(path, data_only=True)["평가"]
    raters = [ws.cell(row=5, column=c).value for c in range(RATER_FIRST_COL, RATER_LAST_COL + 1)]
    depts, rows = [], []
    for r in range(DEPT_FIRST_ROW, DEPT_LAST_ROW + 1):
        depts.append(ws.cell(row=r, column=NAME_COL).value)
        rows.append([ws.cell(row=r, column=c).value for c in range(RATER_FIRST_COL, RATER_LAST_COL + 1)])
    return pd.DataFrame(rows, index=depts, columns=raters)


def co_occurrence(responses: pd.DataFrame) -> np.ndarray:
    """Fraction of raters (who classified both) placing each pair together."""
    m = responses.to_numpy()
    n = len(responses)
    same = np.zeros((n, n))
    valid = np.zeros((n, n))
    for a in range(responses.shape[1]):
        col = m[:, a]
        labelled = np.array([v is not None for v in col])
        for i in range(n):
            if not labelled[i]:
                continue
            for j in range(n):
                if labelled[j]:
                    valid[i, j] += 1
                    if col[i] == col[j]:
                        same[i, j] += 1
    return np.divide(same, valid, where=valid > 0, out=np.zeros_like(same))


def consensus_linkage(co: np.ndarray) -> np.ndarray:
    distance = 1.0 - co
    np.fill_diagonal(distance, 0.0)
    return linkage(squareform(distance, checks=False), method="average")


def algorithm_clusters(k: int) -> tuple[dict, dict]:
    matrix = pd.read_csv(MATRIX, encoding="utf-8-sig")
    course_cols = [c for c in matrix.columns if c not in METADATA]
    names = list(matrix["department_name_ko"])
    x = matrix[course_cols].to_numpy(float)
    idf = np.log(len(matrix) / x.sum(axis=0))

    def cut(weighted):
        d = pdist(weighted, metric="cosine")
        labels = fcluster(linkage(d, method="average"), t=k, criterion="maxclust")
        return dict(zip(names, labels))

    return cut(x * idf), cut(x)


def aligned_idf_cosine(names) -> np.ndarray:
    """IDF cosine similarity matrix, reordered to the card-sort department order."""
    matrix = pd.read_csv(MATRIX, encoding="utf-8-sig")
    course_cols = [c for c in matrix.columns if c not in METADATA]
    order_names = list(matrix["department_name_ko"])
    x = matrix[course_cols].to_numpy(float)
    idf = np.log(len(matrix) / x.sum(axis=0))
    cos = 1 - squareform(pdist(x * idf, metric="cosine"))
    idx = [order_names.index(n) for n in names]
    return cos[np.ix_(idx, idx)]


def disagreement_pairs(co: np.ndarray, names) -> pd.DataFrame:
    """Rank department pairs by internal (IDF) vs external (expert) divergence.

    divergence = z(idf_cosine) - z(co_classification):
      > 0  internal O / external X  (shared course prep, experts keep apart)
      < 0  internal X / external O  (experts group, course prep diverges)
    """
    cos = aligned_idf_cosine(names)
    n = len(names)
    rows = [
        {"A": names[i], "B": names[j],
         "idf_cosine": round(float(cos[i, j]), 3),
         "co_classification": round(float(co[i, j]), 3)}
        for i, j in itertools.combinations(range(n), 2)
    ]
    pairs = pd.DataFrame(rows)
    zi = (pairs["idf_cosine"] - pairs["idf_cosine"].mean()) / pairs["idf_cosine"].std()
    ze = (pairs["co_classification"] - pairs["co_classification"].mean()) / pairs["co_classification"].std()
    pairs["divergence"] = (zi - ze).round(3)
    pairs["type"] = np.where(pairs["divergence"] > 0, "internalO_externalX", "internalX_externalO")
    pairs = pairs.sort_values("divergence", ascending=False).reset_index(drop=True)
    pairs.to_csv(REPORT_TABLES / "cardsort_disagreement_pairs.csv", index=False, encoding="utf-8-sig")
    return pairs


# --------------------------------------------------------------------------- figures
def fig_heatmap(co: np.ndarray, names, order, clusters) -> None:
    ordered = co[np.ix_(order, order)]
    labels = [names[i] for i in order]
    plt.figure(figsize=(10.5, 9))
    plt.imshow(ordered, cmap="magma", vmin=0, vmax=1)
    plt.colorbar(label="동시분류 비율 (co-classification rate)")
    plt.xticks(range(len(labels)), labels, rotation=80, fontsize=8)
    plt.yticks(range(len(labels)), labels, fontsize=8)
    # draw consensus cluster block boundaries along the ordered axis
    ordered_clusters = [clusters[i] for i in order]
    boundaries = [i for i in range(1, len(order)) if ordered_clusters[i] != ordered_clusters[i - 1]]
    for bpos in boundaries:
        plt.axhline(bpos - 0.5, color="cyan", lw=1.2)
        plt.axvline(bpos - 0.5, color="cyan", lw=1.2)
    plt.title("전문가 동시분류 행렬 (14명, consensus 군집 순서)")
    plt.tight_layout()
    plt.savefig(REPORT_FIGURES / "cardsort_cooccurrence_heatmap.png", dpi=200)
    plt.close()


def fig_dendrogram(z: np.ndarray, names) -> None:
    plt.figure(figsize=(12, 6.5))
    dendrogram(z, labels=names, leaf_rotation=80, leaf_font_size=9, color_threshold=0.0)
    plt.title("전문가 consensus dendrogram (1 - 동시분류율)")
    plt.ylabel("Consensus distance")
    plt.tight_layout()
    plt.savefig(REPORT_FIGURES / "cardsort_consensus_dendrogram.png", dpi=200)
    plt.close()


def fig_agreement(agreement: pd.DataFrame) -> None:
    ks = agreement["k"].tolist()
    x = np.arange(len(ks))
    w = 0.38
    plt.figure(figsize=(7, 4.5))
    plt.bar(x - w / 2, agreement["ari_vs_idf"], w, label="vs IDF (main)", color="#2c7fb8")
    plt.bar(x + w / 2, agreement["ari_vs_binary"], w, label="vs binary (baseline)", color="#bdbdbd")
    for xi, (vi, vb) in enumerate(zip(agreement["ari_vs_idf"], agreement["ari_vs_binary"])):
        plt.text(xi - w / 2, vi + 0.01, f"{vi:.2f}", ha="center", fontsize=9)
        plt.text(xi + w / 2, vb + 0.01, f"{vb:.2f}", ha="center", fontsize=9)
    plt.xticks(x, [f"k={k}" for k in ks])
    plt.ylabel("Adjusted Rand Index (전문가 합의와의 일치도)")
    plt.ylim(0, 1)
    plt.title("알고리즘 군집 vs 전문가 합의 (ARI): k=3 강건 일치, k=5–6 IDF 우위")
    plt.legend()
    plt.tight_layout()
    plt.savefig(REPORT_FIGURES / "cardsort_agreement_bars.png", dpi=200)
    plt.close()


def fig_scatter(co: np.ndarray, names) -> None:
    cos = aligned_idf_cosine(names)
    iu = np.triu_indices(len(names), 1)
    xs, ys = cos[iu], co[iu]
    r = np.corrcoef(xs, ys)[0, 1]
    plt.figure(figsize=(6.5, 6))
    plt.scatter(xs, ys, s=18, alpha=0.5, color="#2c7fb8")
    plt.xlabel("IDF cosine similarity (알고리즘)")
    plt.ylabel("전문가 동시분류 비율")
    plt.title(f"학과 쌍(n=300): 알고리즘 vs 전문가  (r = {r:.2f})")
    plt.plot([0, 1], [0, 1], "--", color="gray", lw=1)
    plt.tight_layout()
    plt.savefig(REPORT_FIGURES / "cardsort_cooccur_vs_idf_scatter.png", dpi=200)
    plt.close()
    return r


# --------------------------------------------------------------------------- tables
def write_tables(co, names, clusters_k4, idf_k4, binary_k4, agreement) -> None:
    pd.DataFrame(
        {
            "department": names,
            "consensus_k4": clusters_k4,
            "idf_k4": [idf_k4[n] for n in names],
            "binary_k4": [binary_k4[n] for n in names],
        }
    ).to_csv(REPORT_TABLES / "cardsort_consensus_assignments.csv", index=False, encoding="utf-8-sig")

    agreement.to_csv(REPORT_TABLES / "cardsort_agreement.csv", index=False, encoding="utf-8-sig")

    n = len(names)
    rows = []
    for i in range(n):
        partners = sorted(((co[i, j], names[j]) for j in range(n) if j != i), reverse=True)
        top_val, top_partner = partners[0]
        rows.append({"department": names[i], "max_co_classification": round(top_val, 3),
                     "closest_partner": top_partner})
    pd.DataFrame(rows).sort_values("max_co_classification").to_csv(
        REPORT_TABLES / "cardsort_department_ambiguity.csv", index=False, encoding="utf-8-sig"
    )


def main() -> None:
    REPORT_TABLES.mkdir(parents=True, exist_ok=True)
    REPORT_FIGURES.mkdir(parents=True, exist_ok=True)
    setup_style()

    path = find_input()
    responses = read_responses(path)
    names = list(responses.index)
    n_raters = responses.shape[1]
    co = co_occurrence(responses)
    z = consensus_linkage(co)
    order = list(leaves_list(z))

    clusters_k4 = fcluster(z, t=K_PRIMARY, criterion="maxclust")
    cluster_map = {i: clusters_k4[i] for i in range(len(names))}

    agreement_rows = []
    for k in (3, 4, 5, 6):
        cons_k = fcluster(z, t=k, criterion="maxclust")
        idf_k, bin_k = algorithm_clusters(k)
        idf_aligned = np.array([idf_k[nm] for nm in names])
        bin_aligned = np.array([bin_k[nm] for nm in names])
        agreement_rows.append(
            {
                "k": k,
                "ari_vs_idf": round(adjusted_rand_score(cons_k, idf_aligned), 3),
                "nmi_vs_idf": round(normalized_mutual_info_score(cons_k, idf_aligned), 3),
                "ari_vs_binary": round(adjusted_rand_score(cons_k, bin_aligned), 3),
                "nmi_vs_binary": round(normalized_mutual_info_score(cons_k, bin_aligned), 3),
            }
        )
    agreement = pd.DataFrame(agreement_rows)

    idf_k4, bin_k4 = algorithm_clusters(K_PRIMARY)
    fig_heatmap(co, names, order, cluster_map)
    fig_dendrogram(z, names)
    fig_agreement(agreement)
    r = fig_scatter(co, names)
    write_tables(co, names, clusters_k4, idf_k4, bin_k4, agreement)
    disagree = disagreement_pairs(co, names)

    print(f"input={path.name}  raters={n_raters}  departments={len(names)}")
    print(f"co-occurrence vs IDF cosine: r={r:.3f}")
    print("\nagreement (consensus vs algorithm):")
    print(agreement.to_string(index=False))
    print("\nconsensus k=4 clusters:")
    for c in sorted(set(clusters_k4)):
        print(f"  CC{c}: " + ", ".join(np.array(names)[clusters_k4 == c]))
    print("\nType 1 (내부 O, 외부 X) — 준비기반 공유, 전문가 분리:")
    for _, row in disagree.head(4).iterrows():
        print(f"  {row['A']}-{row['B']}: idf={row['idf_cosine']}, coclass={row['co_classification']}")
    print("Type 2 (내부 X, 외부 O) — 전문가 묶음, 준비과목 분기:")
    for _, row in disagree.tail(4).iloc[::-1].iterrows():
        print(f"  {row['A']}-{row['B']}: idf={row['idf_cosine']}, coclass={row['co_classification']}")
    print("\nfigures -> results/figures/keep_for_report/cardsort_*.png")
    print("tables  -> results/tables/keep_for_report/cardsort_*.csv")


if __name__ == "__main__":
    main()
