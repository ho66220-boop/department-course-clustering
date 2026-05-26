"""Create core raw-data Excel templates for the 25-department scope."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw"

GREEN = PatternFill("solid", fgColor="E2F0D9")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DEPARTMENTS = [
    ("D01", "건축학과", "공학/설계", 1, "공학과 디자인/계획 사이의 boundary case"),
    ("D02", "국어국문학과", "인문", 0, "인문 계열 기준 학과"),
    ("D03", "심리학과", "사회/인문", 1, "사회과학과 인문 사이의 boundary case"),
    ("D04", "기계공학과", "공학", 0, "수학/물리 기반 공학 기준 학과"),
    ("D05", "화학과", "자연과학", 1, "화학 중심 자연과학 학과"),
    ("D06", "산업공학과", "공학/경영", 1, "공학과 경영/데이터 사이의 boundary case"),
    ("D07", "신소재공학과", "공학", 1, "화학/재료 기반 공학 학과"),
    ("D08", "조선해양공학과", "산업특화공학", 1, "울산 지역 산업 맥락을 반영한 applied engineering boundary case"),
    ("D09", "의예과", "의약/보건", 0, "생명/화학 기반 보건의료 학과"),
    ("D10", "수학과", "자연과학/정량", 1, "통계/컴퓨팅/공학과 연결되는 정량 학과"),
    ("D11", "경영학과", "사회", 0, "상담 비교에서 자주 등장하는 사회계열 학과"),
    ("D12", "약학과", "의약/자연과학", 1, "화학/생명 기반 의약 boundary case"),
    ("D13", "사회학과", "사회", 0, "사회과학 기준 학과"),
    ("D14", "경제학과", "사회/정량", 1, "수학/통계와 연결되는 사회과학 boundary case"),
    ("D15", "생명과학과", "자연과학", 1, "생명 중심 자연과학 학과"),
    ("D16", "미디어커뮤니케이션학과", "사회/인문", 1, "사회과학/인문/미디어 사이의 boundary case"),
    ("D17", "화학공학과", "공학/자연과학", 1, "화학과 공학 사이의 boundary case"),
    ("D18", "디자인학과", "예체능/디자인", 1, "건축/미디어와 연결되는 디자인 boundary case"),
    ("D19", "식품영양학과", "자연과학/보건", 1, "생명/간호/식품 분야와 연결되는 학과"),
    ("D20", "간호학과", "의약/보건", 1, "보건의료 상담 관련성이 높은 학과"),
    ("D21", "전기전자공학과", "공학", 0, "수학/물리/정보 기반 공학 학과"),
    ("D22", "사학과", "인문", 0, "인문 계열 기준 학과"),
    ("D23", "응용통계학과", "자연과학/사회", 1, "수학/경제/데이터 분석과 연결되는 boundary case"),
    ("D24", "컴퓨터공학과", "공학/정보", 1, "정보/수학 기반 공학 학과"),
    ("D25", "자동차공학과", "산업특화공학", 1, "울산 지역 산업 맥락을 반영한 applied engineering boundary case"),
]


def style_sheet(ws, widths: dict[str, int], freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = GREEN
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = BORDER
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 34


def add_note_sheet(wb: Workbook, title: str, rows: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet(title)
    ws.append(["Item", "Note"])
    for row in rows:
        ws.append(row)
    style_sheet(ws, {"A": 28, "B": 110})


def add_list_validation(ws, ranges_to_values: dict[str, str]) -> None:
    for cell_range, values in ranges_to_values.items():
        validation = DataValidation(type="list", formula1=f'"{values}"', allow_blank=True)
        ws.add_data_validation(validation)
        validation.add(cell_range)


def create_departments() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "departments"
    ws.append(["department_id", "department_name_ko", "broad_category", "boundary_flag", "include_in_analysis", "note"])
    for department_id, name, category, boundary_flag, note in DEPARTMENTS:
        ws.append([department_id, name, category, boundary_flag, 1, note])
    style_sheet(ws, {"A": 14, "B": 24, "C": 22, "D": 16, "E": 18, "F": 70})
    add_note_sheet(
        wb,
        "guide",
        [
            ("Purpose", "25개 학과 기준표입니다. 이 목록을 기준으로 CareerNet course vector를 구성합니다."),
            ("Sample type", "25개 학과는 statistically representative sample이 아니라 purposive sample입니다."),
            ("Ulsan boundary cases", "조선해양공학과와 자동차공학과는 울산 지역 산업 맥락을 반영한 exploratory boundary case입니다."),
            ("Metadata", "broad_category, boundary_flag, note는 clustering feature로 사용하지 않습니다."),
        ],
    )
    wb.save(RAW_DIR / "departments_raw.xlsx")


def create_recommended_courses() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "recommended_courses"
    headers = [
        "department_id",
        "department_name_ko",
        "source_type",
        "source_name",
        "source_url_or_file",
        "course_name_original",
        "course_name_standardized",
        "subject_group",
        "recommendation_level_original",
        "binary_value",
        "coding_rule",
        "coding_note",
        "source_access_date",
    ]
    ws.append(headers)
    for department_id, department_name, *_ in DEPARTMENTS:
        ws.append(
            [
                department_id,
                department_name,
                "CareerNet",
                "CareerNet major information",
                "https://www.career.go.kr/",
                "",
                "",
                "",
                "related high-school elective subject",
                "",
                "1 if listed as a related high-school elective subject on CareerNet; otherwise 0",
                "",
                "",
            ]
        )
    for _ in range(100):
        ws.append([""] * len(headers))
    style_sheet(ws, {"A": 14, "B": 24, "C": 16, "D": 28, "E": 42, "F": 32, "G": 28, "H": 16, "I": 34, "J": 14, "K": 70, "L": 42, "M": 18})
    add_list_validation(ws, {"C2:C250": "CareerNet,manual_note", "J2:J250": "1,0"})
    add_note_sheet(
        wb,
        "guide",
        [
            ("Purpose", "CareerNet 학과 관련 고교 선택과목을 long-form으로 기록하는 raw template입니다."),
            ("Binary value 1", "CareerNet에서 해당 학과의 관련 고교 선택과목으로 제시된 경우입니다."),
            ("Binary value 0", "최종 matrix 생성 단계에서 미기재 과목을 0으로 처리합니다. raw long-form에는 보통 1 row만 기록합니다."),
            ("Do not sum", "같은 학과-과목 pair가 반복되어도 값은 1로 유지합니다."),
        ],
    )
    wb.save(RAW_DIR / "recommended_courses_raw.xlsx")


def create_source_register() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "source_register"
    ws.append(["source_id", "source_type", "source_name", "source_url_or_file", "access_date", "note"])
    ws.append(["S01", "CareerNet", "CareerNet major information", "https://www.career.go.kr/", "", "Core source for binary related elective subjects"])
    style_sheet(ws, {"A": 14, "B": 16, "C": 30, "D": 48, "E": 16, "F": 70})
    wb.save(RAW_DIR / "raw_source_register.xlsx")


def create_collection_checklist() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "course_collection_status"
    ws.append(["department_id", "department_name_ko", "CareerNet status", "matched_major_name", "matched_url_or_seq", "note"])
    for department_id, department_name, *_ in DEPARTMENTS:
        ws.append([department_id, department_name, "not_started", "", "", ""])
    style_sheet(ws, {"A": 14, "B": 24, "C": 18, "D": 30, "E": 24, "F": 70})
    add_list_validation(ws, {"C2:C100": "not_started,in_progress,completed,no_match"})
    add_note_sheet(wb, "guide", [("Purpose", "25개 학과별 CareerNet 관련 선택과목 수집 진행상황을 체크합니다.")])
    wb.save(RAW_DIR / "collection_checklist.xlsx")


def main() -> None:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    create_departments()
    create_recommended_courses()
    create_source_register()
    create_collection_checklist()
    print("Raw templates recreated:")
    for path in sorted(RAW_DIR.glob("*.xlsx")):
        print(f"- {path}")


if __name__ == "__main__":
    main()
