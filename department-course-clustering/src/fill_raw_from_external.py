"""Legacy external-source filler.

The current Progress-stage scope should use CareerNet binary course evidence
for 25 departments. This older script is retained only as historical context
for external Excel sources and weighted coding experiments.
"""

from __future__ import annotations

import pathlib
import re

import openpyxl


BASE = pathlib.Path(__file__).resolve().parents[1]
EXTERNAL = BASE / "data" / "external"
RAW = BASE / "data" / "raw"

REP_FILE = EXTERNAL / "2028학년도 계열별 대표 모집단위별 반영과목(권장과목).xlsx"
REG_FILE = EXTERNAL / "2028학년도 권역별 대학별 권장과목(반영과목).xlsx"
SCORE_FILE = EXTERNAL / "지거국 교과 종합 컷 정리.xlsx"
COMPREHENSIVE_SCORE_FILE = EXTERNAL / "종합_교과_컷.xlsx"


ALIASES = {
    "D01": ["건축학과", "건축학부", "건축학전공", "건축공학전공", "건축"],
    "D02": ["국어국문학과", "국어국문"],
    "D03": ["심리학과", "심리"],
    "D04": ["기계공학과", "기계공학부", "기계"],
    "D05": ["화학과", "화학"],
    "D06": ["산업공학과", "산업공학"],
    "D07": ["신소재공학과", "신소재"],
    "D08": ["조선해양공학과", "조선해양공학", "조선해양"],
    "D09": ["의예과", "의학과", "의학부", "의예", "의학"],
    "D10": ["수학과", "수학"],
    "D11": ["경영학과", "경영학부", "경영"],
    "D12": ["약학과", "약학부", "약학대학", "약학"],
    "D13": ["사회학과", "사회학"],
    "D14": ["경제학과", "경제학부", "경제"],
    "D15": ["생명과학과", "생명과학", "생명"],
    "D16": ["미디어커뮤니케이션학과", "미디어커뮤니케이션", "언론정보", "신문방송", "미디어"],
    "D17": ["화학공학과", "화공생명공학", "바이오메디컬화학공학", "에너지신소재화학공학", "화학공학"],
    "D18": ["디자인학과", "시각디자인", "디자인"],
    "D19": ["식품영양학과", "식품영양전공", "식품영양"],
    "D20": ["간호학과", "간호대학", "간호"],
    "D21": ["전기전자공학과", "전자전기공학부", "전자전기컴퓨터공학부", "전기·전자", "전기전자", "전자정보", "정보통신전자"],
    "D22": ["사학과", "국사학과", "역사학과", "사학"],
    "D23": ["응용통계학과", "통계학과", "통계"],
    "D24": ["컴퓨터공학과", "컴퓨터공학부", "컴퓨터과학부", "컴퓨터정보공학부", "컴퓨터"],
}

ADMISSION_ALIASES = {
    "D01": ["건축학과", "건축공학과", "건축학부", "건축학전공", "건축공학전공"],
    "D02": ["국어국문학과"],
    "D03": ["심리학과"],
    "D04": ["기계공학과", "기계공학부"],
    "D05": ["화학과"],
    "D06": ["산업공학과"],
    "D07": ["신소재공학과", "화학신소재학과", "신소재시스템공학과"],
    "D08": ["조선해양공학과"],
    "D09": ["의예과", "의학과", "의학부"],
    "D10": ["수학과"],
    "D11": ["경영학과", "경영학부"],
    "D12": ["약학과", "약학부"],
    "D13": ["사회학과"],
    "D14": ["경제학과", "경제학부"],
    "D15": ["생명과학과", "생명시스템과학부", "생명과학부"],
    "D16": ["미디어커뮤니케이션학과", "신문방송학과", "언론정보학과"],
    "D17": ["화학공학과", "화공생명공학과", "화공생명공학부", "바이오메디컬화학공학과"],
    "D18": ["디자인학과", "시각디자인학과", "산업디자인학과"],
    "D19": ["식품영양학과"],
    "D20": ["간호학과"],
    "D21": ["전기전자공학과", "전자전기공학부", "전기공학과", "전자공학과", "정보통신전자공학부"],
    "D22": ["사학과", "국사학과", "역사학과"],
    "D23": ["응용통계학과", "통계학과", "정보통계학과"],
    "D24": ["컴퓨터공학과", "컴퓨터공학부", "컴퓨터과학부", "컴퓨터정보공학부"],
}

ADMISSION_EXCLUDES = {
    "D04": ["바이오산업기계"],
    "D05": ["화학교육과", "생명환경화학"],
    "D09": ["한의", "치의", "수의"],
    "D10": ["수학교육과"],
    "D14": ["식품자원경제", "농업경제"],
    "D15": ["원예생명", "동물생명", "식물생명", "환경생명"],
}

COURSE_EXCLUDES = ADMISSION_EXCLUDES

COURSE_GROUPS = {
    "국어": "국어",
    "화법과 언어": "국어",
    "독서와 작문": "국어",
    "문학": "국어",
    "수학": "수학",
    "대수": "수학",
    "확률과 통계": "수학",
    "미적분": "수학",
    "미적분Ⅰ": "수학",
    "미적분Ⅱ": "수학",
    "기하": "수학",
    "영어": "영어",
    "사회": "사회",
    "일반사회": "사회",
    "역사": "사회",
    "지리": "사회",
    "윤리": "사회",
    "정치와 법": "사회",
    "경제": "사회",
    "사회·문화": "사회",
    "과학": "과학",
    "물리": "과학",
    "물리학": "과학",
    "화학": "과학",
    "생명과학": "과학",
    "지구과학": "과학",
    "역학과 에너지": "과학",
    "전자기와 양자": "과학",
    "물질과 에너지": "과학",
    "화학 반응의 세계": "과학",
    "세포와 물질대사": "과학",
    "생물의 유전": "과학",
    "정보": "기타",
    "인공지능": "기타",
    "체육": "기타",
    "미술": "기타",
    "음악": "기타",
}

COURSE_TERMS = sorted(COURSE_GROUPS, key=len, reverse=True)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def parse_score(value: object) -> float | None:
    if value is None:
        return None
    text = clean_text(value)
    if not text or text in {"-", "None"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).replace("·", "").replace("/", "")


def load_departments() -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(RAW / "departments_raw.xlsx", data_only=True)
    ws = wb["departments"]
    departments = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            departments.append(
                {
                    "id": row[0],
                    "name": row[1],
                    "cat": row[2],
                    "boundary": row[3],
                    "include": row[4],
                    "note": row[5],
                }
            )
    return departments


def match_department(unit_name: object, departments: list[dict[str, object]]) -> tuple[dict[str, object] | None, str]:
    normalized_unit = normalize(unit_name)
    hits = []
    for department in departments:
        department_id = department["id"]
        if any(normalize(excluded) in normalized_unit for excluded in COURSE_EXCLUDES.get(department_id, [])):
            continue
        for alias in ALIASES[department["id"]]:
            normalized_alias = normalize(alias)
            if normalized_alias and normalized_alias in normalized_unit:
                hits.append((len(normalized_alias), department))
                break
    if not hits:
        return None, "none"
    hits.sort(key=lambda item: item[0], reverse=True)
    return hits[0][1], "high" if hits[0][0] >= 4 else "medium"


def match_admission_department(unit_name: object, departments: list[dict[str, object]]) -> tuple[dict[str, object] | None, str]:
    normalized_unit = normalize(unit_name)
    hits = []
    for department in departments:
        department_id = department["id"]
        if any(normalize(excluded) in normalized_unit for excluded in ADMISSION_EXCLUDES.get(department_id, [])):
            continue
        for alias in ADMISSION_ALIASES[department_id]:
            normalized_alias = normalize(alias)
            if normalized_alias and normalized_alias in normalized_unit:
                hits.append((len(normalized_alias), department))
                break
    if not hits:
        return None, "none"
    hits.sort(key=lambda item: item[0], reverse=True)
    return hits[0][1], "high" if hits[0][0] >= 4 else "medium"


def extract_courses(text: object) -> list[str]:
    value = clean_text(text)
    if not value or value in {"-", "None"}:
        return []
    if "진로 및 적성" in value and not any(
        marker in value for marker in ["국어", "수학", "영어", "사회", "과학", "물리", "화학", "생명", "지구", "확률", "미적분", "기하"]
    ):
        return []

    found = []
    for term in COURSE_TERMS:
        if term in value:
            normalized = "물리학" if term == "물리" else term
            if normalized not in found:
                found.append(normalized)
    return found


def build_recommended_course_rows(departments: list[dict[str, object]]) -> list[list[object]]:
    rows = []

    wb = openpyxl.load_workbook(REG_FILE, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not any(row):
            continue
        university = row[2]
        unit_parts = [clean_text(row[index]) for index in (3, 4) if index < len(row) and clean_text(row[index])]
        unit = " / ".join(unit_parts)
        if not unit:
            continue
        department, confidence = match_department(unit, departments)
        if not department:
            continue
        for level, text, weight in [("핵심과목", row[5], 1.0), ("권장과목", row[6], 0.5)]:
            for course in extract_courses(text):
                rows.append(
                    [
                        department["id"],
                        department["name"],
                        "official_doc",
                        "2028학년도 권역별 대학별 권장과목(반영과목)",
                        REG_FILE.name,
                        clean_text(text),
                        course,
                        COURSE_GROUPS.get(course, "기타"),
                        level,
                        weight,
                        "핵심과목=1.0, 권장과목=0.5",
                        f"{university} / {unit} / match={confidence}",
                    ]
                )

    wb = openpyxl.load_workbook(REP_FILE, read_only=True, data_only=True)
    ws = wb["반영과목"]
    headers = []
    top_header = None
    # The source workbook has a sparse tail that makes openpyxl report no
    # max_column in read-only mode. The actual subject columns end at R.
    for column in range(1, 19):
        row3 = ws.cell(3, column).value
        row4 = ws.cell(4, column).value
        if row3:
            top_header = row3
        label = row4 or (row3 if column not in {1, 2} else row3)
        headers.append((column, top_header, label))

    for row in ws.iter_rows(min_row=5, values_only=True):
        if not any(row):
            continue
        unit = row[1]
        department, confidence = match_department(unit, departments)
        if not department:
            continue
        for column, top, label in headers[2:18]:
            value = row[column - 1] if column - 1 < len(row) else None
            if not value or clean_text(value) == "-":
                continue
            course = label
            if course in COURSE_GROUPS:
                rows.append(
                    [
                        department["id"],
                        department["name"],
                        "official_doc",
                        "2028학년도 계열별 대표 모집단위별 반영과목(권장과목)",
                        REP_FILE.name,
                        clean_text(value),
                        "물리학" if course == "물리" else course,
                        COURSE_GROUPS.get(course, str(top)),
                        "대표 모집단위 반영과목",
                        0.5,
                        "대표 모집단위별 반영과목은 보조 근거로 0.5 부여",
                        f"대표모집단위={unit} / match={confidence}",
                    ]
                )

    deduped = []
    seen = set()
    for row in rows:
        key = tuple(row[:10] + [row[11]])
        if key not in seen:
            seen.add(key)
            deduped.append(row)
    deduped.sort(key=lambda item: (item[0], item[4], item[6], item[11]))
    return deduped


def build_admission_score_rows(departments: list[dict[str, object]]) -> list[list[object]]:
    rows = []
    wb = openpyxl.load_workbook(SCORE_FILE, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not any(row):
                continue
            university_original = clean_text(row[0])
            university = sheet_name
            admission_type = clean_text(row[2])
            unit = clean_text(row[5])
            if not unit:
                continue
            department, confidence = match_admission_department(unit, departments)
            if not department:
                continue
            score_candidates = [
                (row[9] if len(row) > 9 else None, "평균/50%컷 등급"),
                (row[10] if len(row) > 10 else None, "70~90%컷 등급②"),
                (row[11] if len(row) > 11 else None, "최저 등급"),
            ]
            for value, score_type in score_candidates:
                score_value = parse_score(value)
                if score_value is None:
                    continue
                rows.append(
                    [
                        university,
                        2025,
                        admission_type,
                        unit,
                        department["id"],
                        department["name"],
                        score_value,
                        "내신등급",
                        score_type,
                        "지거국 교과 종합 컷 정리",
                        SCORE_FILE.name,
                        confidence,
                        f"원본 sheet={sheet_name}; 원본 대학명={university_original or sheet_name}",
                        "",
                    ]
                )
    rows.extend(build_comprehensive_admission_score_rows(departments))
    rows.sort(key=lambda item: (item[0], item[4], item[2], item[8]))
    return rows


def build_comprehensive_admission_score_rows(departments: list[dict[str, object]]) -> list[list[object]]:
    rows = []
    wb = openpyxl.load_workbook(COMPREHENSIVE_SCORE_FILE, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    score_columns = [
        (12, "평균/50%컷 등급", "대학발표 합격 등급"),
        (13, "70~90%컷 등급②", "대학발표 합격 등급"),
        (14, "최저 등급", "대학발표 합격 등급"),
        (16, "평균/50%컷 등급", "대학어디가 발표자료"),
        (17, "70~90%컷 등급②", "대학어디가 발표자료"),
        (18, "최저 등급", "대학어디가 발표자료"),
    ]

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not any(row):
            continue
        university = clean_text(row[5] if len(row) > 5 else "")
        admission_type_group = clean_text(row[0] if len(row) > 0 else "")
        admission_name = clean_text(row[6] if len(row) > 6 else "")
        unit = clean_text(row[7] if len(row) > 7 else "")
        if not university or not unit:
            continue
        department, confidence = match_admission_department(unit, departments)
        if not department:
            continue

        admission_type = " ".join(part for part in [admission_type_group, admission_name] if part)
        extra_note = []
        region = clean_text(row[4] if len(row) > 4 else "")
        university_type = clean_text(row[1] if len(row) > 1 else "")
        theme = clean_text(row[2] if len(row) > 2 else "")
        if region:
            extra_note.append(f"지역={region}")
        if university_type:
            extra_note.append(f"대학구분={university_type}")
        if theme:
            extra_note.append(f"테마={theme}")

        for column_index, score_type, source_section in score_columns:
            value = row[column_index] if len(row) > column_index else None
            score_value = parse_score(value)
            if score_value is None:
                continue
            rows.append(
                [
                    university,
                    2025,
                    admission_type,
                    unit,
                    department["id"],
                    department["name"],
                    score_value,
                    "내신등급",
                    score_type,
                    f"종합_교과_컷 {source_section}",
                    COMPREHENSIVE_SCORE_FILE.name,
                    confidence,
                    "; ".join([f"원본={source_section}"] + extra_note),
                    "2026-05-25",
                ]
            )

    deduped = []
    seen = set()
    for row in rows:
        key = tuple(row[:13])
        if key not in seen:
            seen.add(key)
            deduped.append(row)
    return deduped


def replace_sheet_rows(path: pathlib.Path, sheet_name: str, rows: list[list[object]]) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    for row in rows:
        ws.append(row)
    wb.save(path)


def update_source_register() -> None:
    path = RAW / "raw_source_register.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb["sources"]
    preserved = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] in {"S001", "S002"}:
            preserved.append(list(row[:10]))
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    for row in preserved:
        ws.append(row)
    for row in [
        [
            "S003",
            "official_doc",
            "2028학년도 계열별 대표 모집단위별 반영과목(권장과목)",
            REP_FILE.name,
            "한국대학교육협의회/대입정보포털 계열 자료",
            "manual_excel",
            "2026-05-16",
            "",
            "recommended course evidence",
            "대표 모집단위별 반영과목. 학과 단위가 넓어 보조 근거로 사용",
        ],
        [
            "S004",
            "official_doc",
            "2028학년도 권역별 대학별 권장과목(반영과목)",
            REG_FILE.name,
            "한국대학교육협의회/대입정보포털 권역별 자료",
            "manual_excel",
            "2026-05-16",
            "",
            "recommended course evidence",
            "대학별 핵심과목/권장과목 구분이 있어 course vector 주 근거로 사용",
        ],
        [
            "S005",
            "internal_compiled",
            "지거국 교과 종합 컷 정리",
            SCORE_FILE.name,
            "내부 정리 자료",
            "manual_excel",
            "2026-05-16",
            "",
            "admission score reference",
            "대학/전형/모집단위별 내신 등급 컷. clustering feature가 아닌 해석 변수로만 사용",
        ],
        [
            "S006",
            "internal_compiled",
            "종합_교과_컷",
            COMPREHENSIVE_SCORE_FILE.name,
            "내부 정리 자료",
            "manual_excel",
            "2026-05-25",
            "",
            "admission score reference",
            "전국 대학 교과/종합 전형의 대학발표 및 대학어디가 내신 컷. 기존 지거국 컷을 보강하는 해석 변수로 사용",
        ],
    ]:
        ws.append(row)
    wb.save(path)


def update_collection_checklist(course_rows: list[list[object]], score_rows: list[list[object]]) -> None:
    path = RAW / "collection_checklist.xlsx"
    wb = openpyxl.load_workbook(path)

    course_ids = {row[0] for row in course_rows}
    course_count = {department_id: 0 for department_id in course_ids}
    for row in course_rows:
        course_count[row[0]] += 1

    ws = wb["course_collection_status"]
    for row_index in range(2, ws.max_row + 1):
        department_id = ws.cell(row_index, 1).value
        if not department_id:
            continue
        has_data = department_id in course_ids
        ws.cell(row_index, 3).value = "완료" if has_data else "미확인"
        ws.cell(row_index, 4).value = "미수집"
        ws.cell(row_index, 5).value = "완료" if has_data else "미확인"
        ws.cell(row_index, 6).value = "자동코딩 검토필요" if has_data else "자료부족"
        ws.cell(row_index, 7).value = (
            f"GitHub external 2개 권장과목 파일 자동 매칭; rows={course_count.get(department_id, 0)}"
            if has_data
            else "원본 파일에서 자동 매칭 실패; 수동 확인 필요"
        )

    score_index: dict[tuple[object, object], dict[str, object]] = {}
    for row in score_rows:
        key = (row[0], row[4])
        score_index.setdefault(key, {"avg": False, "cut": False, "n": 0, "units": set()})
        score_index[key]["n"] += 1
        score_index[key]["units"].add(row[3])
        if row[8] == "평균/50%컷 등급":
            score_index[key]["avg"] = True
        if row[8] == "70~90%컷 등급②":
            score_index[key]["cut"] = True

    ws = wb["admission_collection_status"]
    for row_index in range(2, ws.max_row + 1):
        university = ws.cell(row_index, 1).value
        department_id = ws.cell(row_index, 2).value
        if not university or not department_id:
            continue
        info = score_index.get((university, department_id))
        if info:
            units = "; ".join(sorted(info["units"]))[:120]
            ws.cell(row_index, 4).value = "있음"
            ws.cell(row_index, 5).value = "완료" if info["avg"] else "없음"
            ws.cell(row_index, 6).value = "완료" if info["cut"] else "없음"
            ws.cell(row_index, 7).value = "아니오"
            ws.cell(row_index, 8).value = f"자동 매칭 rows={info['n']}; units={units}"
        else:
            ws.cell(row_index, 4).value = "미확인/없음"
            ws.cell(row_index, 5).value = "미수집"
            ws.cell(row_index, 6).value = "미수집"
            ws.cell(row_index, 7).value = "필요"
            ws.cell(row_index, 8).value = "원본 컷 파일에서 자동 매칭 실패 또는 해당 모집단위 없음"

    try:
        wb.save(path)
    except PermissionError:
        print(f"warning: could not save {path}; close the workbook and rerun to update checklist")


def main() -> None:
    departments = load_departments()
    course_rows = build_recommended_course_rows(departments)
    score_rows = build_admission_score_rows(departments)

    replace_sheet_rows(RAW / "recommended_courses_raw.xlsx", "recommended_courses", course_rows)
    replace_sheet_rows(RAW / "admission_scores_raw.xlsx", "admission_scores", score_rows)
    update_source_register()
    update_collection_checklist(course_rows, score_rows)

    print(f"recommended_course_rows={len(course_rows)}")
    print(f"admission_score_rows={len(score_rows)}")
    print(f"course_depts_matched={len({row[0] for row in course_rows})}")
    print(f"admission_depts_matched={len({row[4] for row in score_rows})}")


if __name__ == "__main__":
    main()
