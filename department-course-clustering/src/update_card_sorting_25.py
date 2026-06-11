# -*- coding: utf-8 -*-
"""
Update the card-sorting workbooks from 24 to 25 departments by adding
자동차공학과 (Automotive Engineering) as department #25.

The 25-department scope (incl. 자동차공학과 as an Ulsan industry boundary case)
is the analysis scope used everywhere else in the project, but the original
card-sorting templates were still 24 departments. This script appends the 25th
department at 평가 row 30 and extends every dependent formula, range, data
validation, co-occurrence matrix, pair-data table and QA sheet accordingly.

Run:
    python src/update_card_sorting_25.py
"""
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

CARD_DIR = Path(__file__).resolve().parents[1] / "templates" / "card_sorting"
DIST = CARD_DIR / "CardSorting_distribution_individual_response.xlsx"
MASTER = CARD_DIR / "CardSorting_master_improved.xlsx"

NEW_DEPT = "자동차공학과"
NEW_DEPT_CLASS = "공학"
NEW_DEPT_DESC = "자동차 설계, 동력·구동 시스템"

# 평가 layout constants
FIRST_DEPT_ROW = 6          # dept 1 lives in row 6
N_OLD = 24
N_NEW = 25
NEW_DEPT_ROW = FIRST_DEPT_ROW + N_NEW - 1   # = 30


def cp_style(src, dst):
    """Copy visual style from one cell to another."""
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def eval_row(dept_index):
    """1-based dept index -> its row in the 평가 sheet."""
    return FIRST_DEPT_ROW + dept_index - 1


# --------------------------------------------------------------------------
# 평가 sheet (present in both files)
# --------------------------------------------------------------------------
def update_eval(ws, input_cols, has_response_count, completion_row):
    # 1) add the 25th department row
    ws.cell(row=NEW_DEPT_ROW, column=1, value=N_NEW)
    ws.cell(row=NEW_DEPT_ROW, column=2, value=NEW_DEPT)
    for col in (1, 2):
        cp_style(ws.cell(row=29, column=col), ws.cell(row=NEW_DEPT_ROW, column=col))
    # input/memo cells: copy style from row 29
    for col_letter in input_cols:
        c_idx = openpyxl.utils.column_index_from_string(col_letter)
        cp_style(ws.cell(row=29, column=c_idx), ws.cell(row=NEW_DEPT_ROW, column=c_idx))

    # 2) per-respondent response-count column R (master only)
    if has_response_count:
        src = ws["R29"]
        dst = ws[f"R{NEW_DEPT_ROW}"]
        dst.value = f"=COUNTA(C{NEW_DEPT_ROW}:P{NEW_DEPT_ROW})"
        cp_style(src, dst)

    # 3) completion row: extend range to row 30 and divide by 25
    for col_letter in input_cols:
        cell = ws[f"{col_letter}{completion_row}"]
        if cell.value is not None:
            cell.value = f"=COUNTA({col_letter}6:{col_letter}{NEW_DEPT_ROW})/{N_NEW}"

    # 4) data validation ranges 6:29 -> 6:30
    for dv in list(ws.data_validations.dataValidation):
        new_ranges = []
        for rng in str(dv.sqref).split():
            new_ranges.append(rng.replace("29", str(NEW_DEPT_ROW)))
        dv.sqref = " ".join(new_ranges)


def patch_text(ws, coord, old, new):
    cell = ws[coord]
    if isinstance(cell.value, str) and old in cell.value:
        cell.value = cell.value.replace(old, new)


# --------------------------------------------------------------------------
# 학과상세_참고용 (both files) – append reference row
# --------------------------------------------------------------------------
def update_detail(ws):
    new_row = ws.max_row + 1
    vals = [N_NEW, NEW_DEPT, NEW_DEPT_CLASS, NEW_DEPT_DESC]
    for col, v in enumerate(vals, start=1):
        c = ws.cell(row=new_row, column=col, value=v)
        cp_style(ws.cell(row=ws.max_row - 1 if ws.max_row - 1 >= 5 else 5, column=col), c)
    return new_row


# --------------------------------------------------------------------------
# 집계분석 (master) – grow 24x24 -> 25x25 (add row 29 + column Z)
# --------------------------------------------------------------------------
def cooccurrence_formula(i, j):
    """i = row dept (1..25), j = col dept (1..25)."""
    if i == j:
        return '="-"'
    ri, rj = eval_row(i), eval_row(j)
    a = f"평가!C{ri}:P{ri}"
    b = f"평가!C{rj}:P{rj}"
    return (f'=IFERROR(SUMPRODUCT(({a}<>"")*({b}<>"")*({a}={b}))'
            f'/SUMPRODUCT(({a}<>"")*({b}<>"")),"")')


def update_cooccurrence(ws):
    # header for new column Z (col 26) and new row-header A29
    z_col = 1 + N_NEW  # 26
    hdr = ws.cell(row=4, column=z_col, value=NEW_DEPT)
    cp_style(ws.cell(row=4, column=z_col - 1), hdr)            # copy from Y4
    rhdr = ws.cell(row=4 + N_NEW, column=1, value=NEW_DEPT)    # A29
    cp_style(ws.cell(row=3 + N_NEW, column=1), rhdr)           # copy from A28

    style_src = ws["C5"]  # an existing data cell
    # new column Z: i = 1..25, j = 25
    for i in range(1, N_NEW + 1):
        cell = ws.cell(row=4 + i, column=z_col)
        cell.value = cooccurrence_formula(i, N_NEW)
        cp_style(style_src, cell)
    # new row 29: i = 25, j = 1..24 (Z already set above)
    for j in range(1, N_OLD + 1):
        cell = ws.cell(row=4 + N_NEW, column=1 + j)
        cell.value = cooccurrence_formula(N_NEW, j)
        cp_style(style_src, cell)
    # match column width of Z to Y
    src_dim = ws.column_dimensions[get_column_letter(z_col - 1)]
    ws.column_dimensions[get_column_letter(z_col)].width = src_dim.width


# --------------------------------------------------------------------------
# 분석용_쌍데이터 (master) – append 24 new pairs (k, 25)
# --------------------------------------------------------------------------
def update_pairdata(ws, dept_names):
    # existing rows: 4..279 (header in row 3). pair_id last = 276.
    first_new_row = ws.max_row + 1   # 280
    last_pair_id = N_OLD * (N_OLD - 1) // 2   # 276
    style_row = ws.max_row           # 279, template for styling
    rb = NEW_DEPT_ROW                # row of dept 25 in 평가
    for k in range(1, N_OLD + 1):
        r = first_new_row + (k - 1)
        pid = last_pair_id + k
        ra = eval_row(k)
        a = f"'평가'!C{ra}:P{ra}"
        b = f"'평가'!C{rb}:P{rb}"
        values = {
            "A": pid,
            "B": dept_names[k - 1],
            "C": NEW_DEPT,
            "D": ra,
            "E": rb,
            "F": f'=SUMPRODUCT(({a}<>"")*({b}<>""))',
            "G": f'=SUMPRODUCT(({a}<>"")*({b}<>"")*({a}={b}))',
            "H": f'=IF(F{r}=0,"",G{r}/F{r})',
        }
        for col_letter, v in values.items():
            cell = ws[f"{col_letter}{r}"]
            cell.value = v
            cp_style(ws[f"{col_letter}{style_row}"], cell)


# --------------------------------------------------------------------------
# Response_QA (master) – extend respondent rows + add Group 25 column
# --------------------------------------------------------------------------
def update_response_qa(ws):
    ad_col = 5 + N_NEW  # group columns start at F(=6) for group 1 -> group 25 = col 30 (AD)
    # header
    hdr = ws.cell(row=1, column=ad_col, value=f"Group {N_NEW}")
    cp_style(ws.cell(row=1, column=ad_col - 1), hdr)
    last_group_letter = get_column_letter(ad_col)  # AD

    for r in range(2, 2 + 14):  # 14 respondents -> rows 2..15
        eval_col = get_column_letter(3 + (r - 2))  # C..P
        ws[f"B{r}"].value = f"=COUNT('평가'!{eval_col}6:{eval_col}{NEW_DEPT_ROW})"
        ws[f"C{r}"].value = f'=COUNTIF(F{r}:{last_group_letter}{r},">0")'
        ws[f"D{r}"].value = f"=MAX(F{r}:{last_group_letter}{r})"
        ws[f"E{r}"].value = (
            f'=IF(B{r}=0,"no response yet",IF(B{r}<{N_NEW},"missing depts",'
            f'IF(C{r}=1,"all one group?",IF(C{r}>18,"too many groups?","OK"))))'
        )
        for g in range(1, N_NEW + 1):
            col = 5 + g
            cell = ws.cell(row=r, column=col)
            cell.value = f"=COUNTIF('평가'!${eval_col}$6:${eval_col}${NEW_DEPT_ROW},{g})"
            if col == ad_col:  # newly added column style
                cp_style(ws.cell(row=r, column=ad_col - 1), cell)
    # column width for AD
    ws.column_dimensions[last_group_letter].width = \
        ws.column_dimensions[get_column_letter(ad_col - 1)].width


# --------------------------------------------------------------------------
def process_distribution():
    wb = openpyxl.load_workbook(DIST)
    update_eval(wb["평가"], input_cols=["C", "D"], has_response_count=False, completion_row=32)
    patch_text(wb["평가"], "A3", "1 to 24", "1 to 25")
    # 안내
    patch_text(wb["안내"], "B12", "24개", "25개")
    update_detail(wb["학과상세_참고용"])
    # PreSubmit_Check
    qa = wb["PreSubmit_Check"]
    patch_text(qa, "A2", "24 departments", "25 departments")
    patch_text(qa, "A2", "C6:C29", "C6:C30")
    wb.save(DIST)
    print("saved:", DIST.name)


def process_master():
    wb = openpyxl.load_workbook(MASTER)
    ev = wb["평가"]
    # capture dept name order BEFORE editing
    dept_names = [ev.cell(row=eval_row(k), column=2).value for k in range(1, N_OLD + 1)]
    update_eval(ev, input_cols=[get_column_letter(c) for c in range(3, 17)],  # C..P
                has_response_count=True, completion_row=32)
    patch_text(wb["안내"], "B12", "24개", "25개")
    update_detail(wb["학과상세_참고용"])
    update_cooccurrence(wb["집계분석"])
    update_pairdata(wb["분석용_쌍데이터"], dept_names)
    update_response_qa(wb["Response_QA"])
    # Collection_Guide
    patch_text(wb["Collection_Guide"], "B2", "C6:C29", "C6:C30")
    wb.save(MASTER)
    print("saved:", MASTER.name)


if __name__ == "__main__":
    process_distribution()
    process_master()
    print("done.")
